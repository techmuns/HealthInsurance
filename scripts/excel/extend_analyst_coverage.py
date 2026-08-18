#!/usr/bin/env python3
"""
Analyst-coverage ROW auto-extension (approved by Neha, 2026-08-17).

The Analyst coverage sheet is a list-table: one row per dated broker note, in
date order per company, closed by an Average row. Unlike every other sheet in
the workbook, a new PERIOD here is a new ROW, not a new column — so the period
auto-extension in extend_template_periods.py could never reach it. The effect
was that broker notes newer than the newest template row were fetched, stored
and then dropped on the floor by alignToTuples, because the template named no
cell for them. Niva Bupa sat on a 29-Jan-2026 note while 10-May, 19-Jun and
30-Jul notes were already in the repo.

This script closes that gap the same way: when a real, source-backed, dated
broker note exists that the sheet has no row for, the row is APPENDED into its
company block in date order, and the next schema-map build binds it.

Safety rules (mirroring extend_template_periods.py):
  * Extend ONLY on a real note that carries a date, a source URL and at least
    one of target / price-at-reco. Never a speculative empty row.
  * Value cells (recommendation / price at reco / target) are left EMPTY —
    they are inputs, filled by the sourced data layer on the same run. This
    script never writes a number of its own.
  * Existing rows are never reordered, reworded or renumbered away: every
    original row keeps its own values and its own formula SHAPE (the IPO
    rows' IFERROR form and the literal "Target met" cells are preserved
    exactly). Only the row NUMBERS inside formulas are remapped.
  * Merged company-label ranges follow their block, and each block's Average
    row is re-pointed at that block's true extent.
  * The result is verified against a pre-image before it is saved; any
    mismatch aborts without writing.
"""
from __future__ import annotations

import json
import re
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
TEMPLATE = REPO / "templates" / "niva-bupa-portfolio-review.xlsx"
SNAP = REPO / "src" / "data" / "snapshots"
SHEET = "Analyst coverage"
HEADER_ROW = 3
BROKER_COL, DATE_COL = 3, 4          # C, D
FIRST_VALUE_COL, LAST_VALUE_COL = 5, 10  # E..J

COMPANY_LABELS = {
    "niva bupa": "niva-bupa", "star health": "star-health",
    "icici lombard": "icici-lombard", "go digit": "godigit", "godigit": "godigit",
}


def entity_of(label: str) -> str | None:
    return COMPANY_LABELS.get(re.sub(r"\s+", " ", str(label)).strip().lower())


def norm_broker(s: str) -> str:
    """Broker identity for dedupe: case/punctuation/suffix insensitive."""
    t = re.sub(r"[^a-z0-9 ]", " ", str(s).lower())
    t = re.sub(r"\b(securities|limited|ltd|capital|research|institutional|equities|bnp|paribas)\b", " ", t)
    return re.sub(r"\s+", " ", t).strip()


def parse_date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v or "").strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]{3})[a-z]*,?\s+(\d{4})$", s)  # "30 Jul, 2026"
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%d %b %Y").date().isoformat()
        except ValueError:
            return None
    return None


# ── Candidate notes from the sourced data layer ──────────────────────────────
def load_json(name: str):
    try:
        return json.loads((SNAP / f"{name}.json").read_text())
    except Exception:
        return {}


def candidate_notes() -> list[dict]:
    """Every real, dated, source-backed broker note the data layer holds."""
    out: list[dict] = []
    for r in (load_json("analyst-coverage-snapshot") or {}).get("data", []) or []:
        d = parse_date(r.get("report_date"))
        if not (d and r.get("company_id") and r.get("broker") and r.get("source_url")):
            continue
        if r.get("target_price") is None and r.get("price_at_reco") is None:
            continue
        out.append({"entity": r["company_id"], "broker": str(r["broker"]).strip(), "date": d})
    # Moneycontrol street feed — focal insurer only; it carries the freshest notes.
    st = load_json("street-analyst-snapshot") or {}
    ent = (st.get("_meta") or {}).get("company_id")
    for r in st.get("reports", []) or []:
        d = parse_date(r.get("report_date"))
        if not (d and ent and r.get("brokerage") and r.get("source_url")):
            continue
        if r.get("target_price") is None:
            continue
        out.append({"entity": ent, "broker": str(r["brokerage"]).strip(), "date": d})
    return out


# ── Sheet structure ──────────────────────────────────────────────────────────
class Block:
    def __init__(self, entity, label_row, data_rows, avg_row):
        self.entity, self.label_row = entity, label_row
        self.data_rows, self.avg_row = data_rows, avg_row


def parse_blocks(ws) -> list[Block]:
    blocks, cur_entity, cur_label_row, rows = [], None, None, []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        lbl = ws.cell(r, 2).value
        ent = entity_of(lbl) if lbl else None
        if ent:
            cur_entity, cur_label_row = ent, r
        broker = ws.cell(r, BROKER_COL).value
        if not broker:
            continue
        if str(broker).strip().lower() == "average":
            if cur_entity and rows:
                blocks.append(Block(cur_entity, cur_label_row, rows, r))
            rows, cur_entity, cur_label_row = [], None, None
            continue
        rows.append(r)
    if cur_entity and rows:
        blocks.append(Block(cur_entity, cur_label_row, rows, None))
    return blocks


def formula_shape(v, row: int) -> str:
    """A formula with its OWN row number blanked, so a shifted row still
    compares equal to its pre-image."""
    if not isinstance(v, str) or not v.startswith("="):
        return "" if v is None else str(v)
    return re.sub(rf"(?<![0-9]){row}(?![0-9])", "#", v)


def remap_formula(v, row_map: dict[int, int]) -> object:
    """Rewrite every row reference through row_map. Handles self-references
    (=H4/G4-1), cross-references (=D14) and ranges (=AVERAGE(F4:F17)) alike."""
    if not isinstance(v, str) or not v.startswith("="):
        return v
    def sub(m):
        col, row = m.group(1), int(m.group(2))
        return f"{col}{row_map.get(row, row)}"
    return re.sub(r"(\$?[A-J])(\d+)\b", sub, v)



def date_ref_target(ws, row: int):
    """For a date cell that references another row (=D14), the broker named on
    that row — a stable identity across renumbering. None for a literal date."""
    v = ws.cell(row, DATE_COL).value
    if not (isinstance(v, str) and v.startswith("=")):
        return None
    m = re.match(r"^=\$?D(\d+)$", v.strip())
    if not m:
        return v
    return str(ws.cell(int(m.group(1)), BROKER_COL).value or "").strip()


def snapshot_rows(ws, blocks) -> dict:
    """Pre-image: per company, each data row's identity, values and formula
    SHAPE. Compared after the rewrite so nothing can be silently altered."""
    pre = {}
    for b in blocks:
        rows = []
        for r in b.data_rows:
            rows.append({
                "broker": ws.cell(r, BROKER_COL).value,
                "date": parse_date(ws.cell(r, DATE_COL).value),
                # A date cell may BORROW another row's date (=D14 on the IPO rows).
                # Record what it points AT, not the row number, so a mis-remapped
                # reference is caught even though the number legitimately changes.
                "date_ref": date_ref_target(ws, r),
                "vals": [ws.cell(r, c).value if not isinstance(ws.cell(r, c).value, str)
                         or not str(ws.cell(r, c).value).startswith("=") else None
                         for c in range(FIRST_VALUE_COL, LAST_VALUE_COL + 1)],
                "shapes": [formula_shape(ws.cell(r, c).value, r)
                           for c in range(FIRST_VALUE_COL, LAST_VALUE_COL + 1)],
            })
        pre[b.entity] = rows
    return pre


def plan(ws, blocks, notes) -> dict[int, list[dict]]:
    """position (insert BEFORE this original row) -> notes to place there."""
    by_entity: dict[str, list[dict]] = {}
    for n in notes:
        by_entity.setdefault(n["entity"], []).append(n)
    positions: dict[int, list[dict]] = {}
    for b in blocks:
        existing = [(r, norm_broker(ws.cell(r, BROKER_COL).value), parse_date(ws.cell(r, DATE_COL).value))
                    for r in b.data_rows]
        present = {(nb, d) for _, nb, d in existing if d}
        dated = [(r, d) for r, _, d in existing if d]
        first_undated = next((r for r, _, d in existing if not d), None)
        end_pos = b.avg_row or (b.data_rows[-1] + 1)
        seen = set()
        fresh = []
        for n in sorted(by_entity.get(b.entity, []), key=lambda x: x["date"], reverse=True):
            key = (norm_broker(n["broker"]), n["date"])
            if key in present or key in seen:
                continue
            seen.add(key)
            fresh.append(n)
        for n in fresh:
            pos = next((r for r, d in dated if d < n["date"]), first_undated or end_pos)
            positions.setdefault(pos, []).append(n)
    for p in positions:
        positions[p].sort(key=lambda x: x["date"], reverse=True)
    return positions


def main() -> int:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb[SHEET]
    blocks = parse_blocks(ws)
    if not blocks:
        print("Analyst coverage: no company blocks parsed — sheet shape changed; refusing to touch it.")
        return 1
    pre = snapshot_rows(ws, blocks)
    positions = plan(ws, blocks, candidate_notes())
    if not positions:
        print("Analyst coverage: no new dated broker notes — template already names every one.")
        return 0

    old_merges = [str(m) for m in ws.merged_cells.ranges]
    for m in old_merges:
        ws.unmerge_cells(m)

    # Apply insertions bottom-up so the lower positions keep their indices.
    for pos in sorted(positions, key=lambda p: -p):
        ws.insert_rows(pos, len(positions[pos]))

    def shift(r: int) -> int:
        return r + sum(len(v) for p, v in positions.items() if p <= r)
    row_map = {r: shift(r) for r in range(1, ws.max_row + 2)}

    # Every formula in the sheet still names pre-insert rows; remap them all.
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = remap_formula(cell.value, row_map)

    # Fill the new rows: identity + date only. Values stay empty for the data
    # layer to fill; the upside formulas are cloned from the block's donor row.
    added = []
    for pos, notes in positions.items():
        donor_old = next(b.data_rows[0] for b in blocks if b.data_rows[0] <= pos <= (b.avg_row or b.data_rows[-1] + 1))
        donor = row_map[donor_old]
        for i, n in enumerate(notes):
            r = row_map[pos] - len(notes) + i
            for c in range(2, LAST_VALUE_COL + 1):
                src, dst = ws.cell(donor, c), ws.cell(r, c)
                if src.has_style:
                    dst.font, dst.border = copy(src.font), copy(src.border)
                    dst.fill, dst.alignment = copy(src.fill), copy(src.alignment)
                    dst.number_format = src.number_format
            ws.cell(r, BROKER_COL).value = n["broker"]
            ws.cell(r, DATE_COL).value = datetime.fromisoformat(n["date"])
            ws.cell(r, 6).value = ws.cell(donor, 6).value  # CMP — same static column value
            for c in (9, 10):  # I, J upside formulas, re-pointed at this row
                f = ws.cell(donor, c).value
                if isinstance(f, str) and f.startswith("="):
                    ws.cell(r, c).value = re.sub(rf"(?<![0-9]){donor}(?![0-9])", str(r), f)
            added.append((n["entity"], n["broker"], n["date"], r))

    # Re-anchor company labels + merges, and re-point each Average row.
    new_blocks = parse_blocks(ws)
    for b in new_blocks:
        first, last = b.data_rows[0], b.data_rows[-1]
        label = next((ws.cell(r, 2).value for r in b.data_rows if ws.cell(r, 2).value), None)
        for r in b.data_rows:
            if r != first:
                ws.cell(r, 2).value = None
        ws.cell(first, 2).value = label
        end = last
        for m in old_merges:
            mm = re.match(r"^B(\d+):B(\d+)$", m)
            if mm and int(mm.group(1)) in row_map and row_map[int(mm.group(1))] in range(first, last + 2):
                end = min(max(row_map[int(mm.group(2))], first), last)
        if end > first:
            ws.merge_cells(start_row=first, start_column=2, end_row=end, end_column=2)
        if b.avg_row:
            for c, col in ((6, "F"), (7, "G"), (8, "H")):
                ws.cell(b.avg_row, c).value = f"=AVERAGE({col}{first}:{col}{last})"
    return finish(wb, ws, pre, added)


def finish(wb, ws, pre, added) -> int:
    """Verify the rewrite against the pre-image, then save. Any drift aborts."""
    post_blocks = parse_blocks(ws)
    post = snapshot_rows(ws, post_blocks)
    problems = []
    for entity, want in pre.items():
        got = post.get(entity)
        if got is None:
            problems.append(f"{entity}: block disappeared")
            continue
        it = iter(got)
        for w in want:
            for g in it:  # originals must still appear, in their original order
                if (g["broker"] == w["broker"] and g["date"] == w["date"]
                        and g["date_ref"] == w["date_ref"]
                        and g["vals"] == w["vals"] and g["shapes"] == w["shapes"]):
                    break
            else:
                problems.append(f"{entity}: lost or altered row {w['broker']} {w['date']}")
                break
        n_new = len(got) - len(want)
        n_exp = sum(1 for a in added if a[0] == entity)
        if n_new != n_exp:
            problems.append(f"{entity}: {n_new} rows added, expected {n_exp}")
    for b in post_blocks:
        if not b.avg_row:
            continue
        want = f"=AVERAGE(F{b.data_rows[0]}:F{b.data_rows[-1]})"
        if ws.cell(b.avg_row, 6).value != want:
            problems.append(f"{b.entity}: Average range {ws.cell(b.avg_row, 6).value} != {want}")
        for r in b.data_rows:
            if not ws.cell(r, BROKER_COL).value:
                problems.append(f"{b.entity}: blank broker at row {r}")
    if problems:
        print("Analyst coverage: ABORTED, template NOT written:")
        for p in problems:
            print(f"  ✗ {p}")
        return 1
    wb.save(TEMPLATE)
    print(f"Analyst coverage: appended {len(added)} broker note row(s) to the template.")
    for entity, broker, date, row in sorted(added, key=lambda a: (a[0], a[2]), reverse=True):
        print(f"  + {entity:<14} {date}  {broker}  (row {row})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
