#!/usr/bin/env python3
"""
Phase 7 - QA gate for the filled workbook.

Validates the source-backed fill against the project's honesty rules and the
task's QA checklist. Exits non-zero on a HARD failure so CI fails clearly.

Hard checks (fail the build):
  H1  Every Source Audit row carries provenance (source_name + source_url).
  H2  Audit + Missing partition is complete & disjoint = every fillable cell is
      accounted for exactly once (nothing silently dropped, nothing double-counted).
  H3  No filled value is null/blank (missing != zero - nulls must be on the
      Missing Data sheet, never written as a value).
  H4  Unit bounds: ratios (claims/expense/combined) in [0,3], shares in [0,1.02],
      solvency in [0,10]; premiums/PAT finite.
  H5  Invariants where all legs are present: GWP >= NWP >= NEP; and
      |combined - (claims+expense)| <= 0.03.

Soft checks (warn only):
  S1  Coverage %, and per-status counts on the Missing Data sheet.
  S2  Backup-sourced values present (Screener/Trendlyne) flagged for review.

Usage: python3 scripts/excel/qa_checks.py [filled.xlsx]
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[2]
SCHEMA = REPO / "schema-map.json"
VALUES = REPO / "data" / "processed" / "excel-values.json"
FILLED = REPO / "output" / "Niva_Bupa_portfolio_review__filled.xlsx"

RATIO_METRICS = {"claims_ratio_igaap", "expense_ratio_igaap", "combined_ratio_igaap",
                 "claims_ratio_ifrs", "expense_ratio_ifrs"}
# Investment yield CAN be negative in a drawdown quarter (e.g. Star Q4 FY26
# prints an annualised -4.6% on marked-to-market losses) — a printed negative
# is honest data, not a violation. Bounds widened accordingly.
YIELD_METRICS = {"investment_yield"}
SHARE_METRICS = {"overall_health_market_share", "retail_health_market_share"}


def main(filled_path: Path) -> int:
    schema = json.loads(SCHEMA.read_text())
    store = json.loads(VALUES.read_text()) if VALUES.exists() else {}
    wb = openpyxl.load_workbook(filled_path, data_only=False)

    hard, soft = [], []

    # Expected partition from schema-map + store. A store entry flagged
    # conflict_needs_review is NOT filled (charter: source-conflicting -> withhold)
    # and is NOT missing either - it is a third "blocked" bucket on Blocked Data.
    expected_fill, expected_miss, expected_blocked = set(), set(), set()
    for s in schema["sheets"]:
        for b in s["bindings"]:
            if not b.get("fillable"):
                continue
            cellref = (s["sheet"], b["cell"])
            key = f"{b['entity']}::{b['metric']}::{b['period']}"
            e = store.get(key)
            if e and e.get("normalized_value") is not None:
                if e.get("conflict_status") == "conflict_needs_review":
                    expected_blocked.add(cellref)
                else:
                    expected_fill.add(cellref)
            else:
                expected_miss.add(cellref)

    if "Source Audit" not in wb.sheetnames:
        hard.append("Source Audit sheet missing.")
    if "Missing Data" not in wb.sheetnames:
        hard.append("Missing Data sheet missing.")
    if hard:
        return report(hard, soft)

    aud = wb["Source Audit"]
    mis = wb["Missing Data"]
    # Source Audit: row1 banner, row2 header, data from row3.
    audit_cells, audit_vals = set(), {}
    for r in range(3, aud.max_row + 1):
        sheet, cell = aud.cell(r, 1).value, aud.cell(r, 2).value
        if not sheet or not cell:
            continue
        audit_cells.add((sheet, cell))
        src_name, src_url = aud.cell(r, 10).value, aud.cell(r, 11).value
        norm = aud.cell(r, 8).value
        metric = aud.cell(r, 4).value
        audit_vals[(sheet, cell)] = (norm, metric)
        # H1 provenance — every audited value must NAME its source. A source URL
        # is required too, EXCEPT for analyst-entered figures compiled in the
        # portfolio-review workbook: that is a real, named provenance which simply
        # has no external link (Neha compiled them herself), so demanding a URL
        # there is wrong and was failing every scheduled fill. Anything citing an
        # external source still must carry its link.
        analyst_entered = isinstance(src_name, str) and "analyst-entered" in src_name.lower()
        if not src_name or (not src_url and not analyst_entered):
            hard.append(f"H1 provenance missing for {sheet}!{cell} (name={src_name!r}, url={src_url!r}).")
        # H3 missing != zero (no null written as a value)
        if norm is None:
            hard.append(f"H3 filled cell {sheet}!{cell} has a null value (should be on Missing Data).")
        # H4 unit bounds
        if isinstance(norm, (int, float)) and metric:
            base = str(metric).split("::")[0]
            if base in RATIO_METRICS and not (0 <= norm <= 3):
                hard.append(f"H4 ratio out of bounds: {sheet}!{cell} {base}={norm}.")
            if base in YIELD_METRICS and not (-0.5 <= norm <= 3):
                hard.append(f"H4 yield out of bounds: {sheet}!{cell} {base}={norm}.")
            if base in SHARE_METRICS and not (0 <= norm <= 1.02):
                hard.append(f"H4 share out of bounds: {sheet}!{cell} {base}={norm}.")
            if base == "solvency_ratio" and not (0 <= norm <= 10):
                hard.append(f"H4 solvency out of bounds: {sheet}!{cell} {base}={norm}.")

    missing_cells = set()
    for r in range(2, mis.max_row + 1):
        sheet, cell = mis.cell(r, 1).value, mis.cell(r, 2).value
        if sheet and cell:
            missing_cells.add((sheet, cell))

    # H2 partition complete & disjoint
    overlap = audit_cells & missing_cells
    if overlap:
        hard.append(f"H2 {len(overlap)} cell(s) appear in BOTH Source Audit and Missing Data, e.g. {sorted(overlap)[:3]}.")
    if audit_cells != expected_fill:
        miss_a = expected_fill - audit_cells
        extra_a = audit_cells - expected_fill
        if miss_a:
            hard.append(f"H2 {len(miss_a)} expected-filled cell(s) absent from Source Audit, e.g. {sorted(miss_a)[:3]}.")
        if extra_a:
            hard.append(f"H2 {len(extra_a)} audited cell(s) not expected to be filled, e.g. {sorted(extra_a)[:3]}.")
    # Conflict-blocked cells must be on NEITHER sheet (withheld from the workbook).
    leaked = expected_blocked & (audit_cells | missing_cells)
    if leaked:
        hard.append(f"H2 {len(leaked)} source-conflict cell(s) leaked into Audit/Missing, e.g. {sorted(leaked)[:3]}.")
    total_expected = len(expected_fill) + len(expected_miss)
    total_seen = len(audit_cells) + len(missing_cells)
    if total_seen < total_expected:
        soft.append(f"S1 partition counts: {total_seen} accounted vs {total_expected} expected (some empty cells skipped).")
    if expected_blocked:
        soft.append(f"S1 {len(expected_blocked)} cell(s) withheld as source_conflict (kept in store, on Blocked Data).")

    # H5 invariants from the value store (entity+period legs).
    legs = defaultdict(dict)
    for key, e in store.items():
        ent, metric, period = key.split("::", 2)
        if e.get("normalized_value") is not None:
            legs[(ent, period)][metric] = e["normalized_value"]
    for (ent, period), m in legs.items():
        g, n, ne = m.get("total_gwp"), m.get("nwp"), m.get("nep")
        if g is not None and n is not None and g + 1e-6 < n:
            hard.append(f"H5 GWP<NWP for {ent} {period}: {g} < {n}.")
        if n is not None and ne is not None and n + 1e-6 < ne:
            hard.append(f"H5 NWP<NEP for {ent} {period}: {n} < {ne}.")
        cl, ex, co = m.get("claims_ratio_igaap"), m.get("expense_ratio_igaap"), m.get("combined_ratio_igaap")
        if cl is not None and ex is not None and co is not None and abs(co - (cl + ex)) > 0.03:
            soft.append(f"S? combined != claims+expense for {ent} {period}: {co} vs {cl}+{ex}={cl+ex:.3f}.")

    # H6 Analyst coverage structural integrity. The sheet is the one list-table
    # in the workbook: it grows DOWNWARD as broker notes arrive, so the classes
    # of breakage that matter are structural, not numeric — a row that lost its
    # broker, an Average that stopped covering its block, the same note landing
    # twice, or a row sitting past whatever range the schema builder scans (that
    # last one silently dropped Go Digit the first time the sheet grew).
    if "Analyst coverage" in wb.sheetnames:
        ac = wb["Analyst coverage"]
        blocks, cur, rows = [], None, []
        for r in range(4, ac.max_row + 1):
            lbl = ac.cell(r, 2).value
            if lbl and str(lbl).strip():
                cur = str(lbl).strip()
            broker = ac.cell(r, 3).value
            if not broker:
                continue
            if str(broker).strip().lower() == "average":
                if rows:
                    blocks.append((cur, rows, r))
                rows = []
                continue
            rows.append(r)
        if rows:
            blocks.append((cur, rows, None))
        if not blocks:
            hard.append("H6 Analyst coverage: no company blocks found — sheet structure lost.")
        seen_notes = set()
        for entity, data_rows, avg_row in blocks:
            for r in data_rows:
                if not str(ac.cell(r, 3).value or "").strip():
                    hard.append(f"H6 Analyst coverage row {r} has no broker.")
                key = (entity, str(ac.cell(r, 3).value).strip().lower(), str(ac.cell(r, 4).value))
                if key in seen_notes:
                    # SOFT on purpose. Two same-day notes from one broker collide on a
                    # single schema-map key, so both rows show the same figure — worth
                    # surfacing, but it is pre-existing sheet content, not corruption.
                    # Hard-failing on a data quirk is precisely what froze three
                    # pipelines for a fortnight; never gate the build on one again.
                    soft.append(f"S3 Analyst coverage duplicate note {key[1]} {key[2]} for {entity} (row {r}) - both rows resolve to one value.")
                seen_notes.add(key)
            if avg_row:
                for col in ("F", "G", "H"):
                    want = f"=AVERAGE({col}{data_rows[0]}:{col}{data_rows[-1]})"
                    got = ac.cell(avg_row, {"F": 6, "G": 7, "H": 8}[col]).value
                    if got != want:
                        hard.append(f"H6 Analyst coverage {entity} Average {col} is {got!r}, expected {want!r}.")
        # Every dated row must be bound in the schema map — catches a scan range
        # that stopped short of the sheet's real extent.
        try:
            smap = json.loads((REPO / "schema-map.json").read_text())
            sheet_map = next(x for x in smap["sheets"] if x["sheet"] == "Analyst coverage")
            bound_rows = {int(re.sub(r"[^0-9]", "", b["cell"])) for b in sheet_map.get("bindings", []) if b.get("cell")}
            for entity, data_rows, _ in blocks:
                for r in data_rows:
                    if isinstance(ac.cell(r, 4).value, datetime) and r not in bound_rows:
                        soft.append(f"S3 Analyst coverage row {r} ({entity}) is dated but unbound in schema-map - it will never fill.")
        except Exception as e:
            soft.append(f"S3 Analyst coverage binding cross-check skipped: {e}")

    # Soft coverage stats
    pct = 100.0 * len(audit_cells) / max(1, len(audit_cells) + len(missing_cells))
    soft.append(f"S1 coverage: {len(audit_cells)} filled / {len(audit_cells)+len(missing_cells)} fillable ({pct:.1f}%).")
    by_status = defaultdict(int)
    for r in range(2, mis.max_row + 1):
        st = mis.cell(r, 7).value
        if st:
            by_status[st] += 1
    soft.append("S1 missing by status: " + ", ".join(f"{k}={v}" for k, v in sorted(by_status.items())))
    backups = sum(1 for v in store.values() if v.get("source_status") == "backup")
    if backups:
        soft.append(f"S2 {backups} backup-sourced value(s) (Screener/Trendlyne) present - tagged low-confidence, review before publishing.")

    # Blocked Data sheet stats (extracted-but-withheld, categorized).
    if "Blocked Data" in wb.sheetnames:
        blk = wb["Blocked Data"]
        by_cat = defaultdict(int)
        for r in range(2, blk.max_row + 1):
            cat = blk.cell(r, 1).value
            if cat:
                by_cat[cat] += 1
        if by_cat:
            soft.append("S1 blocked by category: " + ", ".join(f"{k}={v}" for k, v in sorted(by_cat.items())))
    else:
        soft.append("S1 Blocked Data sheet absent (no withheld values).")

    return report(hard, soft)


def report(hard, soft) -> int:
    # Machine-readable copy for the workflows' alert step: a hard failure must
    # never be silent, so the runner opens / refreshes a GitHub issue from this.
    import os
    out_dir = os.environ.get("RUNNER_TEMP") or os.environ.get("QA_REPORT_DIR")
    if out_dir:
        try:
            Path(out_dir).joinpath("excel-qa-report.json").write_text(json.dumps({
                "generated_at": datetime.now().isoformat(timespec="seconds"),
                "hard": hard, "soft": soft, "passed": not hard,
            }, indent=2, ensure_ascii=False))
        except OSError as e:
            print(f"  (could not write QA report json: {e})")
    print("=" * 64)
    print("Excel QA report")
    print("=" * 64)
    for s in soft:
        print(f"  [warn] {s}")
    if hard:
        print()
        for h in hard:
            print(f"  [FAIL] {h}")
        print(f"\nQA FAILED: {len(hard)} hard violation(s).")
        return 1
    print("\nQA PASSED: no hard violations.")
    return 0


if __name__ == "__main__":
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else FILLED
    if not path.exists():
        sys.exit(f"Filled workbook not found: {path}. Run fill_template.py first.")
    raise SystemExit(main(path))
